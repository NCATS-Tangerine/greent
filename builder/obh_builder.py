from greent.rosetta import Rosetta
from greent import node_types
from greent.util import LoggingUtil
from greent.graph_components import KNode, KEdge
from builder.buildmain import run
from builder.question import LabeledID
from multiprocessing import Pool
from greent.export import BufferedWriter
from functools import partial
from openpyxl import load_workbook
import logging, time

logger = LoggingUtil.init_logging(__name__, level=logging.DEBUG)

class ObesityHubBuilder(object):

    reference_chrom_versions = {
        'GRCh37': {
            'p1': {
                '1': '10',
                '2': '11',
                '3': '11',
                '4': '11',
                '5': '9',
                '6': '11',
                '7': '13',
                '8': '10',
                '9': '11',
                '10': '10',
                '11': '9',
                '12': '11',
                '13': '10',
                '14': '8',
                '15': '9',
                '16': '9',
                '17': '10',
                '18': '9',
                '19': '9',
                '20': '10',
                '21': '8',
                '22': '10',
                '23': '10',
                '24': '9',
                'X': '10',
                'Y': '9'
            }
        },
        'GRCh38': {
            'p1': {
                '1': '11',
                '2': '12',
                '3': '12',
                '4': '12',
                '5': '10',
                '6': '12',
                '7': '14',
                '8': '11',
                '9': '12',
                '10': '11',
                '11': '10',
                '12': '12',
                '13': '11',
                '14': '9',
                '15': '10',
                '16': '10',
                '17': '11',
                '18': '10',
                '19': '10',
                '20': '11',
                '21': '9',
                '22': '11',
                '23': '11',
                '24': '10',
                'X': '11',
                'Y': '10'
            }
        }
    }

    reference_prefixes = {
        'GRCh37': 'NC_0000',
        'GRCh38': 'NC_0000'
    }

    def __init__(self, rosetta, debug=False):
        self.rosetta = rosetta
        self.concept_model = rosetta.type_graph.concept_model
        if debug:
            logger.setLevel(logging.DEBUG)

    def get_metabolites_from_file(self, metabolites_file):
        metabolite_nodes = []
        metabolite_file_names = {}
        try:
            wb = load_workbook(filename=metabolites_file, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                if row[5].value != '#N/A':
                    m_id = f'PUBCHEM:{row[5].value}'
                elif row[7].value != '#N/A':
                    m_id = f'HMDB:{row[7].value}'
                elif row[6].value != '#N/A':
                    m_id = f'KEGG.COMPOUND:{row[6].value}'
                else:
                    continue

                m_label = row[0].value
                m_filename = f'{row[8].value}'
                # temporary workaround for incorrect names
                #m_filename = f'{row[8].value}_scale'
                #m_filename = m_filename[0:32]
                metabolite_file_names[m_id] = m_filename

                new_metabolite_node = KNode(m_id, name=m_label, type=node_types.CHEMICAL_SUBSTANCE)
                metabolite_nodes.append(new_metabolite_node)

        except (IOError, KeyError) as e:
            logger.warning(f'metabolites_file ({metabolites_file}) could not be loaded or parsed: {e}')

        return metabolite_nodes, metabolite_file_names

    def create_gwas_graph(self, source_nodes, gwas_file_names, gwas_file_directory, p_value_cutoff, reference_genome='GRCh37', reference_patch='p1'):
        variants_processed = 0
        predicate = LabeledID(identifier=f'RO:0002609', label=f'related_to')
        pool = Pool(processes=8)

        for source_node in source_nodes:
            filepath = f'{gwas_file_directory}/{gwas_file_names[source_node.id]}'
            identifiers, p_values = self.get_hgvs_identifiers_from_vcf(filepath, p_value_cutoff, reference_genome, reference_patch)
            if len(identifiers) > 0:
                self.rosetta.synonymizer.synonymize(source_node)
                with BufferedWriter(self.rosetta) as writer:
                    writer.write_node(source_node)

                for seq_var_id in identifiers:
                    p_value = p_values.get(seq_var_id.identifier)
                    self.write_new_association(source_node, seq_var_id, node_types.SEQUENCE_VARIANT, predicate, p_value)
                    
                partial_run_one = partial(find_connections, node_types.SEQUENCE_VARIANT, node_types.GENE)
                pool.map(partial_run_one, identifiers)

                partial_run_one = partial(find_connections, node_types.SEQUENCE_VARIANT, node_types.DISEASE_OR_PHENOTYPIC_FEATURE)
                pool.map(partial_run_one, identifiers)

                variants_processed += len(identifiers)

        pool.close()
        pool.join()

        logger.info(f'create_gwas_graph complete - {variants_processed} significant variants found and processed.')

    def write_new_association(self, source_node, associated_node_id, associated_node_type, predicate, p_value):
        
        associated_node = KNode(associated_node_id.identifier, name=associated_node_id.label, type=associated_node_type)
        self.rosetta.synonymizer.synonymize(associated_node)

        if self.concept_model:
            standard_predicate = self.concept_model.standardize_relationship(predicate)
        else:
            logger.warning('obesity builder: concept_model was missing, predicate standardization failed')
            standard_predicate = predicate

        props={'p_value': p_value}
        ctime = time.time()
        new_edge = KEdge(source_id=source_node.id,
                     target_id=associated_node.id,
                     provided_by='Obesity_Hub',
                     ctime=ctime,
                     original_predicate=predicate,
                     standard_predicate=standard_predicate,
                     input_id=source_node.id,
                     publications=None,
                     url=None,
                     properties=props)
        with BufferedWriter(self.rosetta) as writer:
            writer.write_node(associated_node)
            writer.write_edge(new_edge)

    def get_hgvs_identifiers_from_vcf(self, filename, p_value_cutoff, reference_genome, reference_patch):
        new_ids = []
        corresponding_p_values = {}
        try:
            with open(filename) as f:
                headers = next(f).split()
                line_counter = 0
                try:
                    pval_index = headers.index('PVALUE')
                    chrom_index = headers.index('CHROM')
                    pos_index = headers.index('POS')
                    ref_index = headers.index('REF')
                    alt_index = headers.index('ALT')
                except ValueError:
                    logger.warning(f'Error reading file headers for {filename}')
                for line in f:
                    try:
                        line_counter += 1
                        data = line.split()
                        chromosome = data[chrom_index]
                        position = data[pos_index]
                        ref_allele = data[ref_index]
                        alt_allele = data[alt_index]
                        p_value = data[pval_index]
                    
                        if (p_value != 'NA') and (float(p_value) <= p_value_cutoff):
                            hgvs = self.convert_vcf_to_hgvs(reference_genome, reference_patch, chromosome, position, ref_allele, alt_allele)
                            if hgvs:
                                curie_hgvs = f'HGVS:{hgvs}'
                                new_ids.append(LabeledID(identifier=curie_hgvs, label=f'Variant(hgvs): {hgvs}'))
                                corresponding_p_values[curie_hgvs] = p_value

                    except (IndexError, ValueError) as e:
                        logger.warning(f'Error reading file {filename}, on line {line_counter}: {e}')

        except IOError:
            logger.warning(f'Could not open file: {filename}')

        return new_ids, corresponding_p_values

    def convert_vcf_to_hgvs(self, reference_genome, reference_patch, chromosome, position, ref_allele, alt_allele):
        try:
            ref_chrom_version = self.reference_chrom_versions[reference_genome][reference_patch][chromosome]
            ref_prefix = self.reference_prefixes[reference_genome]

        except KeyError:
            logger.warning(f'Reference chromosome and/or version not found: {reference_genome}.{reference_patch},{chromosome}')
            return ''

        if chromosome == 'X':
            chromosome = '23'
        elif chromosome == 'Y':
            chromosome = '24'
        elif len(chromosome) == 1:
            ref_prefix = f'{ref_prefix}0'

        len_ref =  len(ref_allele) 
        len_alt = len(alt_allele)
            
        if len_alt == 1:
            # deletions
            if alt_allele == '.' or not alt_allele:
                if len_ref is 1:
                    variation = f'{position}del'
                else:
                    variation = f'{position}_{int(position)+len_ref}del'
            # substitutions
            elif len_ref == 1:      
                variation = f'{position}{ref_allele}>{alt_allele}'
            # more deletions
            elif len_ref == 2 and (ref_allele[0] == alt_allele[0]):
                variation = f'{int(position)+1}del'
            elif len_ref > 2 and (ref_allele[0] == alt_allele[0]):
                variation = f'{int(position)+1}_{int(position)+len_ref}del'

        # insertions
        elif (len_alt > len_ref) and (len_ref == 1) and (ref_allele[0] == alt_allele[0]):
            variation = f'{position}_{int(position) + 1}ins{alt_allele[1:]}'

        else:
            logger.warning(f'Format of variant not recognized for hgvs conversion: {ref_allele} to {alt_allele}')
            return ''

        hgvs = f'{ref_prefix}{chromosome}.{ref_chrom_version}:g.{variation}'
        return hgvs

def find_connections(input_type, output_type, identifier):
    path = f'{input_type},{output_type}'
    run(path,identifier.label,identifier.identifier,None,None,None,'greent.conf')

if __name__=='__main__':
    
    obh = ObesityHubBuilder(Rosetta(), debug=True)

    #metabolites_file = '/projects/sequence_analysis/vol1/obesity_hub/metabolomics/files_for_using_metabolomics_data/SOL_metabolomics_info_10202017.xlsx'
    #gwas_directory = '/projects/sequence_analysis/vol1/obesity_hub/metabolomics/aggregate_results'
    #metabolites_file = './sample_metabolites.xlsx'
    #gwas_directory = '.'
    #p_value_cutoff = .000001

    #create a graph with just one node / file
    #pa_id = 'EFO:0003940'
    #pa_node = KNode(pa_id, name='Physical Activity', type=node_types.DISEASE_OR_PHENOTYPIC_FEATURE)
    #associated_nodes = [pa_node]
    #associated_file_names = {pa_id:''}
    #gwas_directory = '/projects/sequence_analysis/vol1/obesity_hub/PA'
    #obh.create_gwas_graph(associated_nodes, associated_file_names, gwas_directory, p_value_cutoff)

    #create a graph with a file of many nodes
    #metabolite_nodes, metabolite_file_names = obh.get_metabolites_from_file(metabolites_file)
    #obh.create_gwas_graph(metabolite_nodes, metabolite_file_names, gwas_directory, p_value_cutoff)
   
    # generic example / shows reference param
    #obh.create_gwas_graph(associated_nodes, associated_node_file_names, gwas_directory, p_value_cutoff, reference_genome='GRCh38')