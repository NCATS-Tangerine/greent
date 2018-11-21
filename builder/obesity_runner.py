from greent.rosetta import Rosetta
from greent import node_types
from greent.graph_components import KNode, KEdge
from builder.buildmain import run
from builder.question import LabeledID
from multiprocessing import Pool
from greent.export import BufferedWriter
from functools import partial
from openpyxl import load_workbook
import time

class ObesityHub(object):

    reference_chrom_versions = {
        'GRCh37': {
            'p1': {
                '1': '10',
                '2': '11',
                '3': '11',
                '4': '11',
                '5': '9',
                '6': '11',
                '7': '13',
                '8': '10',
                '9': '11',
                '10': '10',
                '11': '9',
                '12': '11',
                '13': '10',
                '14': '8',
                '15': '9',
                '16': '9',
                '17': '10',
                '18': '9',
                '19': '9',
                '20': '10',
                '21': '8',
                '22': '10',
                '23': '10',
                '24': '9',
                'X': '10',
                'Y': '9'
            }
        },
        'GRCh38': {
            'p1': {
                '1': '11',
                '2': '12',
                '3': '12',
                '4': '12',
                '5': '10',
                '6': '12',
                '7': '14',
                '8': '11',
                '9': '12',
                '10': '11',
                '11': '10',
                '12': '12',
                '13': '11',
                '14': '9',
                '15': '10',
                '16': '10',
                '17': '11',
                '18': '10',
                '19': '10',
                '20': '11',
                '21': '9',
                '22': '11',
                '23': '11',
                '24': '10',
                'X': '11',
                'Y': '10'
            }
        }
    }

    reference_prefixes = {
        'GRCh37': 'NC_0000'
    }

    def __init__(self, rosetta):
        self.rosetta = rosetta

    def get_hgvs_identifiers_from_vcf(self, filename, p_value_cutoff, reference_genome, reference_patch):
        new_ids = []
        corresponding_p_values = {}
        try:
            with open(filename) as f:
                headers = next(f).split()
                if ('PVALUE' in headers):
                    pval_index = headers.index('PVALUE')
                else:
                    print(f'Invalid file format: {filename}')
                    return 0
                for line in f:
                    data = line.split()
                    if len(data) < pval_index-1:
                        continue
                    chromosome = data[0]
                    position = data[1]
                    ref_allele = data[3]
                    alt_allele = data[4]
                    p_value = data[pval_index]
                    try:
                        if float(p_value) <= p_value_cutoff:
                            hgvs = self.convert_vcf_to_hgvs(reference_genome, reference_patch, chromosome, position, ref_allele, alt_allele)
                            if hgvs:
                                new_ids.append(LabeledID(identifier=f'HGVS:{hgvs}', label=f'Variant(hgvs): {hgvs}'))
                                corresponding_p_values[hgvs] = p_value

                    except ValueError:
                        continue
                        # should we log this? p value wasn't a float

        except IOError:
            print(f'Could not open file: {filename}')

        return new_ids, corresponding_p_values

    def convert_vcf_to_hgvs(self, reference_genome, reference_patch, chromosome, position, ref_allele, alt_allele):
        if (reference_genome in self.reference_prefixes and reference_genome in self.reference_chrom_versions
            and reference_patch in self.reference_chrom_versions[reference_genome] 
            and chromosome in self.reference_chrom_versions[reference_genome][reference_patch]):
            
            ref_chrom_version = self.reference_chrom_versions[reference_genome][reference_patch][chromosome]
            ref_prefix = self.reference_prefixes[reference_genome]
            if chromosome == 'X':
                chromosome = '23'
            elif chromosome == 'Y':
                chromosome = '24'
            elif len(chromosome) == 1:
                ref_prefix = f'{ref_prefix}0'

            len_ref =  len(ref_allele) 
            len_alt = len(alt_allele)
            
            if len_alt == 1:
                # deletions
                if alt_allele == '.' or not alt_allele:
                    if len_ref is 1:
                        variation = f'{position}del'
                    else:
                        variation = f'{position}_{int(position)+len_ref}del'
                # substitutions
                elif len_ref == 1:      
                    variation = f'{position}{ref_allele}>{alt_allele}'
                # more deletions
                elif len_ref == 2 and (ref_allele[0] == alt_allele[0]):
                    variation = f'{int(position)+1}del'
                elif len_ref > 2 and (ref_allele[0] == alt_allele[0]):
                    variation = f'{int(position)+1}_{int(position)+len_ref}del'

            # insertions
            elif (len_alt > len_ref) and (ref_allele[0] == alt_allele[0]):
                variation = f'{position}_{int(position) + 1}ins{alt_allele[1:]}'

            else:
                print(f'Format of variant not recognized for hgvs conversion: {ref_allele} to {alt_allele}')
                return ''

            hgvs = f'{ref_prefix}{chromosome}.{ref_chrom_version}:g.{variation}'
            #print(hgvs)
            return hgvs

        else:
            print(f'Reference chromosome version not found: {reference_genome}.{reference_patch},{chromosome}')
            return ''

    def create_obesity_graph(self, metabolites_file, gwas_directory, p_value_cutoff, reference_genome, reference_patch='p1'):
        pool = Pool(processes=4)
        variants_processed = 0
        try:
            wb = load_workbook(filename=metabolites_file, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2):
                m_label = row[0].value
                m_filename = f'{gwas_directory}/{row[8].value}'
                m_id = ''
                if row[5].value != '#N/A':
                    m_id = f'PUBCHEM:{row[5].value}'
                elif row[7].value != '#N/A':
                    m_id = f'HMDB:{row[7].value}'
                elif row[6].value != '#N/A':
                    m_id = f'KEGG.COMPOUND:{row[6].value}'
                else:
                    continue

                identifiers, p_values = self.get_hgvs_identifiers_from_vcf(m_filename, p_value_cutoff, reference_genome, reference_patch)
                if len(identifiers) > 0:
                    metabolite_node = KNode(m_id, name=m_label, type=node_types.CHEMICAL_SUBSTANCE)
                    self.rosetta.synonymizer.synonymize(metabolite_node)
                    with BufferedWriter(self.rosetta) as writer:
                        writer.write_node(metabolite_node)

                    for seq_id in identifiers:
                        self.write_experimental_edge(metabolite_node, seq_id, node_types.SEQUENCE_VARIANT, p_values, time.time())
                    
                    partial_run_one = partial(find_connections, node_types.SEQUENCE_VARIANT, node_types.GENE)
                    pool.map(partial_run_one, identifiers)

                    partial_run_one = partial(find_connections, node_types.SEQUENCE_VARIANT, node_types.DISEASE_OR_PHENOTYPIC_FEATURE)
                    pool.map(partial_run_one, identifiers)

                    variants_processed += len(identifiers)

        except IOError:
            print(f'metabolites_file ({metabolites_file}) could not be loaded as xlsx')

        pool.close()
        pool.join()

        print(f'{variants_processed} significant variants found and processed')

    def write_experimental_edge(self, source_node, associated_node_id, associated_node_type, p_values, ctime):
        
        associated_node = KNode(associated_node_id.identifier, type=associated_node_type)
        associated_node.name = associated_node_id.label
        self.rosetta.synonymizer.synonymize(associated_node)
        predicate = LabeledID(identifier=f'OBH:experimental_association', label=f'experimental_association')
        props={'p_value': p_values.get(associated_node_id.identifier)}
        new_edge = KEdge(source_id=source_node.id,
                     target_id=associated_node.id,
                     provided_by='Obesity_Hub',
                     ctime=ctime,
                     original_predicate=predicate,
                     standard_predicate=predicate,
                     input_id=source_node.id,
                     publications=None,
                     url=None,
                     properties=props)
        with BufferedWriter(self.rosetta) as writer:
            writer.write_node(associated_node)
            writer.write_edge(new_edge)

def find_connections(input_type, output_type, identifier):
        path = f'{input_type},{output_type}'
        #print(f'{identifier.identifier} - {path}')
        run(path,identifier.label,identifier.identifier,None,None,None,'greent.conf')

if __name__=='__main__':
    #metabolites_file = '/projects/sequence_analysis/vol1/obesity_hub/metabolomics/files_for_using_metabolomics_data/SOL_metabolomics_info_10202017.xlsx'
    #gwas_directory = '/projects/sequence_analysis/vol1/obesity_hub/metabolomics/aggregate_results/'
    metabolites_file = './sample_metabolites.xlsx'
    gwas_directory = '.'
    obh = ObesityHub(Rosetta())
    obh.create_obesity_graph(metabolites_file, gwas_directory, .000001, 'GRCh37')

